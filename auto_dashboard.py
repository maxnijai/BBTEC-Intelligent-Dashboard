# BBTEC Intelligent Dashboard — Auto-Update via Google Drive + GitHub Actions

import io
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from math import atan2, cos, radians, sin, sqrt
from pathlib import Path

import csv
import urllib.request
import urllib.error
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

GDRIVE_FOLDER_ID = "188bv1FhdU2A64wjJjkntkKFIGZG18nMJ"

# NASA FIRMS - VIIRS NOAA-20 (NRT) 24h - no API key required for public CSV
# Coverage: Thailand bounding box (lat 5-21, lon 97-106)
FIRMS_URL = (
    "https://firms.modaps.eosdis.nasa.gov/api/country/csv/"
    "2f5e32d9c58a0b96e61e43c9de47bd38/"  # public map key
    "VIIRS_SNPP_NRT/THA/1"               # Thailand, 1 day
)
# Fallback: MODIS NRT (broader coverage, lower resolution)
FIRMS_MODIS_URL = (
    "https://firms.modaps.eosdis.nasa.gov/api/country/csv/"
    "2f5e32d9c58a0b96e61e43c9de47bd38/"
    "MODIS_NRT/THA/1"
)
# Northern Thailand bounding box
NORTH_LAT_MIN, NORTH_LAT_MAX = 15.0, 21.0
NORTH_LON_MIN, NORTH_LON_MAX = 97.0, 102.5

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "dashboard_output"
SERVICE_ACCOUNT_FILE = SCRIPT_DIR / "service_account.json"
LAST_TRACKER = SCRIPT_DIR / "last_processed.txt"
TPL_FILE = SCRIPT_DIR / "templates" / "tpl_component.txt"

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

AF_ORDER = [
    "6) Within SLA",
    "5) OverSLA : < 1 day",
    "4) OverSLA : < 3 days",
    "3) OverSLA : < 7 days",
    "2) OverSLA : < 30 days",
    "1) OverSLA : > 30 days",
]


def haversine(lat1, lon1, lat2, lon2):
    r = 6371
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
    return r * 2 * atan2(sqrt(a), sqrt(1 - a))


def get_drive_svc():
    if not SERVICE_ACCOUNT_FILE.exists():
        raise FileNotFoundError(f"service_account.json not found: {SERVICE_ACCOUNT_FILE}")

    creds = service_account.Credentials.from_service_account_file(
        str(SERVICE_ACCOUNT_FILE),
        scopes=SCOPES,
    )
    return build("drive", "v3", credentials=creds)


def find_latest_gdrive():
    svc = get_drive_svc()
    res = svc.files().list(
        q=(
            f"'{GDRIVE_FOLDER_ID}' in parents and "
            "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
            "and trashed=false"
        ),
        orderBy="modifiedTime desc",
        pageSize=1,
        fields="files(id,name,modifiedTime)",
    ).execute()

    files = res.get("files", [])
    if not files:
        return None, 0, None

    f = files[0]
    mt = datetime.fromisoformat(f["modifiedTime"].replace("Z", "+00:00")).timestamp()
    print("  GDrive:", f["name"])
    return f, mt, svc


def download_gdrive(svc, fid, fname):
    OUTPUT_DIR.mkdir(exist_ok=True)
    fp = OUTPUT_DIR / fname

    req = svc.files().get_media(fileId=fid)
    fh = io.BytesIO()
    dl = MediaIoBaseDownload(fh, req)

    done = False
    while not done:
        _, done = dl.next_chunk()

    fh.seek(0)
    with open(fp, "wb") as f:
        f.write(fh.read())

    return str(fp)


def find_best_file():
    gf, _, svc = find_latest_gdrive()
    if gf:
        return download_gdrive(svc, gf["id"], gf["name"]), gf["name"]
    return None, None


def process_excel(fp):
    print("  Processing:", os.path.basename(fp))
    wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)

    if "data" not in wb.sheetnames:
        wb.close()
        raise KeyError("Sheet 'data' not found in Excel file")

    ws = wb["data"]
    all_rows = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is None:
            continue

        reg = str(row[11]) if row[11] else ""
        if reg not in ("NOR1", "NOR2"):
            continue

        owner = str(row[12]) if row[12] else ""
        if "UN-NPMECT-RF" in owner:
            continue

        sev = str(row[7] or "")
        af = str(row[22] or "")
        prov = str(row[36]) if row[36] and row[36] != "None" else ""
        district = str(row[38]) if row[38] and row[38] != "None" else ""
        site = str(row[35] or "")
        tid = str(row[3] or "")
        subject = str(row[8])[:80] if row[8] else ""
        aging = round(float(row[21] or 0), 2)
        sowner = owner.split("BBT-")[1] if "BBT-" in owner else owner
        bookmark = str(row[25] or "")
        cat = str(row[10] or "")
        itime = str(row[0] or "")
        creation = str(row[5] or "")
        cls_full = str(row[19] or "")

        parts = [p.strip() for p in cls_full.split("  ") if p.strip()]
        vendor, problem, root = "Other", "Other", "Other"

        if len(parts) >= 2:
            vp = parts[1]
            for v, k in [("ERICSSON", "Ericsson"), ("NOKIA", "Nokia"), ("HUAWEI", "Huawei")]:
                if v in vp:
                    vendor = k
                    break
            for pt in ["ROUTE SITE DOWN", "SITE UP/DOWN", "SITE DOWN", "CELL DOWN"]:
                if pt in vp:
                    problem = pt
                    break

        if len(parts) >= 3:
            root = parts[2]

        subimpact = 0
        try:
            subimpact = int(float(str(row[13]))) if row[13] and row[13] != "None" else 0
        except Exception:
            pass

        lat = row[23]
        lon = row[24]
        flat, flon, hc = 0, 0, False
        if lat and lat != "None" and lon and lon != "None":
            try:
                flat = round(float(lat), 4)
                flon = round(float(lon), 4)
                hc = True
            except Exception:
                pass

        status = str(row[4] or "")
        ciname = str(row[9] or "")
        target = str(row[6] or "")
        over_sla = str(row[2] or "")
        plan_close = str(row[18]) if row[18] and row[18] != "None" else ""
        ext_sys = str(row[30] or "")

        pri = 2
        now = datetime.now()
        ref_time = datetime(now.year, now.month, now.day, 1, 15, 0)
        if ref_time <= now:
            ref_time = ref_time + timedelta(days=1)

        if target and target != "None":
            try:
                dt = datetime.strptime(target[:19], "%Y-%m-%d %H:%M:%S")
                diff_h = (ref_time - dt).total_seconds() / 3600
                if diff_h > 24:
                    pri = 0
                elif diff_h > 0:
                    pri = 1
                else:
                    pri = 2
            except Exception:
                pass

        all_rows.append(
            {
                "sev": sev,
                "reg": reg,
                "sowner": sowner,
                "vendor": vendor,
                "problem": problem,
                "root": root,
                "af": af,
                "aging": aging,
                "prov": prov,
                "bookmark": bookmark,
                "cat": cat,
                "lat": flat,
                "lon": flon,
                "hc": hc,
                "itime": itime,
                "district": district,
                "subimpact": subimpact,
                "creation": creation,
                "site": site,
                "cls": cls_full,
                "pri": pri,
                "ticket": {
                    "tid": tid,
                    "sev": sev,
                    "reg": reg,
                    "af": af,
                    "aging": aging,
                    "sowner": sowner,
                    "prov": prov,
                    "dist": district,
                    "status": status,
                    "cat": cat,
                    "problem": problem,
                    "root": root,
                    "site": site,
                    "ci": ciname,
                    "sub": subimpact,
                    "created": creation[:16],
                    "target": target[:16],
                    "over_sla": over_sla,
                    "plan": plan_close[:16],
                    "ext": ext_sys,
                    "bookmark": bookmark,
                    "lat": flat,
                    "lon": flon,
                    "subj": subject,
                    "cls": cls_full,
                    "pri": pri,
                },
            }
        )

    wb.close()

    if not all_rows:
        print("  No data")
        return None

    def cluster_5km(rows):
        tix = [r for r in rows if r["hc"]]
        if not tix:
            return []

        used = set()
        clusters = []

        for idx in sorted(range(len(tix)), key=lambda i: -tix[i]["aging"]):
            if idx in used:
                continue

            t = tix[idx]
            members = []
            for idx2 in range(len(tix)):
                if idx2 in used:
                    continue
                if haversine(t["lat"], t["lon"], tix[idx2]["lat"], tix[idx2]["lon"]) <= 5:
                    members.append(idx2)
                    used.add(idx2)

            if len(members) >= 2:
                ctix = [tix[j] for j in members]
                sites = list(set(t2["site"] for t2 in ctix if t2["site"]))
                af_c = Counter(t2["af"] for t2 in ctix)
                tkt_det = [
                    {
                        "tid": t2["ticket"]["tid"],
                        "site": t2["site"],
                        "subj": t2["ticket"]["subj"],
                        "aging": t2["aging"],
                        "af": t2["af"],
                        "sev": t2["sev"],
                        "lat": t2["lat"],
                        "lon": t2["lon"],
                        "cls": t2["cls"],
                        "pri": t2["pri"],
                    }
                    for t2 in sorted(ctix, key=lambda x: -x["aging"])
                ]
                clusters.append(
                    {
                        "lat": round(sum(t2["lat"] for t2 in ctix) / len(ctix), 4),
                        "lon": round(sum(t2["lon"] for t2 in ctix) / len(ctix), 4),
                        "count": len(ctix),
                        "sites": len(sites),
                        "site_names": sites[:8],
                        "prov": ctix[0]["prov"],
                        "dist": ctix[0]["district"],
                        "reg": ctix[0]["reg"],
                        "avg": round(sum(t2["aging"] for t2 in ctix) / len(ctix), 2),
                        "max": round(max(t2["aging"] for t2 in ctix), 2),
                        "sowners": list(set(t2["sowner"] for t2 in ctix)),
                        "tkt": tkt_det,
                        **{k: af_c.get(k, 0) for k in AF_ORDER},
                    }
                )

        clusters.sort(key=lambda x: -x["count"])
        return clusters[:30]

    def build_weather(rows):
        hourly = defaultdict(lambda: {"total": 0, "power": 0, "provs": Counter()})
        for r in rows:
            c = r["creation"]
            if c and ("2026-03-11" in c or "2026-03-12" in c):
                hr = c[11:13] if len(c) > 12 else "00"
                key = c[:10] + " " + hr + ":00"
                hourly[key]["total"] += 1
                if "POWER" in r["root"]:
                    hourly[key]["power"] += 1
                if r["prov"]:
                    hourly[key]["provs"][r["prov"]] += 1

        timeline = [
            {
                "time": k,
                "total": d["total"],
                "power": d["power"],
                "provs": ",".join(p + "(" + str(c) + ")" for p, c in d["provs"].most_common(3)),
            }
            for k, d in sorted(hourly.items())
        ]
        peak = sorted(timeline, key=lambda x: -x["total"])[:5]
        t11 = sum(h["total"] for k, h in hourly.items() if "03-11" in k)
        t12 = sum(h["total"] for k, h in hourly.items() if "03-12" in k)
        p11 = sum(h["power"] for k, h in hourly.items() if "03-11" in k)
        p12 = sum(h["power"] for k, h in hourly.items() if "03-12" in k)
        return {
            "timeline": timeline,
            "peak": peak,
            "summary": {
                "total_11mar": t11,
                "total_12mar": t12,
                "power_11mar": p11,
                "power_12mar": p12,
                "peak_hour": peak[0]["time"] if peak else "",
                "peak_count": peak[0]["total"] if peak else 0,
                "storm_total": t11 + t12,
                "storm_power": p11 + p12,
            },
        }

    def fetch_firms_hotspots():
        """Fetch real-time fire hotspots from NASA FIRMS API (VIIRS + MODIS NRT).
        Returns list of dicts: {lat, lon, brightness, frp, acq_date, acq_time, confidence, instrument}
        """
        hotspots = []
        urls = [
            (FIRMS_URL, "VIIRS_SNPP"),
            (FIRMS_MODIS_URL, "MODIS"),
        ]
        for url, instrument in urls:
            try:
                req = urllib.request.Request(
                    url,
                    headers={"User-Agent": "BBTEC-Dashboard/1.0"}
                )
                with urllib.request.urlopen(req, timeout=15) as resp:
                    raw = resp.read().decode("utf-8")
                reader = csv.DictReader(raw.splitlines())
                for row in reader:
                    try:
                        lat = float(row.get("latitude") or row.get("lat") or 0)
                        lon = float(row.get("longitude") or row.get("lon") or 0)
                        # Filter to Northern Thailand only
                        if not (NORTH_LAT_MIN <= lat <= NORTH_LAT_MAX and
                                NORTH_LON_MIN <= lon <= NORTH_LON_MAX):
                            continue
                        brightness = float(row.get("bright_ti4") or row.get("brightness") or 0)
                        frp = float(row.get("frp") or 0)
                        confidence = str(row.get("confidence") or "n")
                        acq_date = str(row.get("acq_date") or "")
                        acq_time = str(row.get("acq_time") or "")
                        hotspots.append({
                            "lat": round(lat, 4),
                            "lon": round(lon, 4),
                            "brightness": round(brightness, 1),
                            "frp": round(frp, 2),
                            "confidence": confidence,
                            "acq_date": acq_date,
                            "acq_time": acq_time,
                            "instrument": instrument,
                        })
                    except (ValueError, KeyError):
                        continue
                print(f"  FIRMS {instrument}: {len(hotspots)} hotspots in Northern Thailand")
            except urllib.error.URLError as e:
                print(f"  FIRMS {instrument} fetch failed: {e}")
            except Exception as e:
                print(f"  FIRMS {instrument} error: {e}")
        return hotspots

    def build_fire(rows, hotspots):
        """Correlate NASA FIRMS hotspots with ticket sites.
        Analyzes proximity, risk levels, and correlation with ticket surge/root cause.
        """
        if not hotspots:
            return None

        # Get unique sites with coordinates
        site_map = {}
        for r in rows:
            if r["hc"] and r["site"]:
                key = r["site"]
                if key not in site_map:
                    site_map[key] = {
                        "site": r["site"],
                        "prov": r["prov"],
                        "reg": r["reg"],
                        "lat": r["lat"],
                        "lon": r["lon"],
                        "tickets": [],
                    }
                site_map[key]["tickets"].append(r)

        # For each site, find nearest hotspot and count within 5km
        site_fire = []
        for site_name, s in site_map.items():
            distances = []
            for hs in hotspots:
                d = haversine(s["lat"], s["lon"], hs["lat"], hs["lon"])
                distances.append((d, hs))
            distances.sort(key=lambda x: x[0])

            if not distances:
                continue

            nearest_km = distances[0][0]
            nearest_hs = distances[0][1]
            count_1km = sum(1 for d, _ in distances if d <= 1)
            count_5km = sum(1 for d, _ in distances if d <= 5)
            count_10km = sum(1 for d, _ in distances if d <= 10)

            # Only include sites with fire within 10km
            if nearest_km > 10:
                continue

            open_tickets = len(s["tickets"])
            over_sla = sum(1 for t in s["tickets"] if t["af"] != "6) Within SLA")
            avg_aging = round(sum(t["aging"] for t in s["tickets"]) / len(s["tickets"]), 1) if s["tickets"] else 0
            has_power_fail = any("POWER" in t["root"] for t in s["tickets"])
            has_hardware = any("HARDWARE" in t["root"].upper() or "EQUIPMENT" in t["root"].upper() for t in s["tickets"])

            # Risk scoring: distance + count + ticket correlation
            risk_score = 0
            if nearest_km < 1:
                risk_score += 50
            elif nearest_km < 3:
                risk_score += 30
            elif nearest_km < 5:
                risk_score += 15
            elif nearest_km < 10:
                risk_score += 5
            risk_score += min(count_5km * 3, 30)  # cap at 30
            if has_power_fail:
                risk_score += 10
            if has_hardware:
                risk_score += 5
            if avg_aging > 10:
                risk_score += 5

            site_fire.append({
                "site": site_name,
                "prov": s["prov"],
                "reg": s["reg"],
                "lat": s["lat"],
                "lon": s["lon"],
                "nearest_km": round(nearest_km, 2),
                "nearest_frp": nearest_hs["frp"],
                "nearest_brightness": nearest_hs["brightness"],
                "count_1km": count_1km,
                "count_5km": count_5km,
                "count_10km": count_10km,
                "open_tickets": open_tickets,
                "over_sla": over_sla,
                "avg_aging": avg_aging,
                "has_power_fail": has_power_fail,
                "has_hardware": has_hardware,
                "risk_score": risk_score,
            })

        # Sort by risk score descending
        site_fire.sort(key=lambda x: -x["risk_score"])

        # Province-level aggregation
        prov_risk = {}
        for s in site_fire:
            p = s["prov"]
            if p not in prov_risk:
                prov_risk[p] = {"sites": 0, "hotspots_5km": 0, "tickets": 0, "critical": 0}
            prov_risk[p]["sites"] += 1
            prov_risk[p]["hotspots_5km"] += s["count_5km"]
            prov_risk[p]["tickets"] += s["open_tickets"]
            if s["nearest_km"] < 1:
                prov_risk[p]["critical"] += 1

        top_prov = max(prov_risk.items(), key=lambda x: x[1]["hotspots_5km"])[0] if prov_risk else "-"

        # Correlation analysis: compare ticket metrics near vs far from fire
        near_tickets = [r for r in rows if r["hc"] and any(
            haversine(r["lat"], r["lon"], hs["lat"], hs["lon"]) <= 5
            for hs in hotspots
        )]
        far_tickets = [r for r in rows if r["hc"] and r not in near_tickets]

        near_avg_aging = round(sum(t["aging"] for t in near_tickets) / len(near_tickets), 1) if near_tickets else 0
        far_avg_aging = round(sum(t["aging"] for t in far_tickets) / len(far_tickets), 1) if far_tickets else 0
        near_power_pct = round(sum(1 for t in near_tickets if "POWER" in t["root"]) / max(len(near_tickets), 1) * 100, 1)
        far_power_pct = round(sum(1 for t in far_tickets if "POWER" in t["root"]) / max(len(far_tickets), 1) * 100, 1)
        near_over_pct = round(sum(1 for t in near_tickets if t["af"] != "6) Within SLA") / max(len(near_tickets), 1) * 100, 1)
        far_over_pct = round(sum(1 for t in far_tickets if t["af"] != "6) Within SLA") / max(len(far_tickets), 1) * 100, 1)

        # Top root causes near fire
        near_roots = Counter(t["root"] for t in near_tickets)
        top_near_roots = [{"name": k, "count": v} for k, v in near_roots.most_common(5)]

        # Hotspot density by province (for map overlay)
        hs_by_prov = Counter()
        for hs in hotspots:
            # Simple province lookup by coordinate bounding box
            for pname, (lat_min, lat_max, lon_min, lon_max) in {
                "เชียงราย": (19.0, 20.5, 99.3, 100.9),
                "เชียงใหม่": (17.9, 20.1, 97.7, 99.5),
                "แม่ฮ่องสอน": (17.5, 20.0, 97.3, 98.2),
                "ลำปาง": (17.8, 19.6, 99.0, 100.3),
                "พะเยา": (18.8, 19.7, 99.9, 100.9),
                "น่าน": (18.3, 20.0, 100.4, 101.5),
                "แพร่": (17.8, 18.9, 99.7, 100.8),
                "ลำพูน": (17.8, 18.7, 98.6, 99.4),
                "ตาก": (15.5, 18.2, 97.4, 99.5),
                "สุโขทัย": (16.7, 17.5, 99.1, 100.0),
                "อุตรดิตถ์": (17.2, 18.8, 99.8, 101.5),
                "พิษณุโลก": (16.3, 17.8, 99.8, 101.2),
                "กำแพงเพชร": (15.6, 16.9, 98.9, 100.2),
                "เพชรบูรณ์": (15.7, 17.8, 100.4, 101.8),
            }.items():
                if lat_min <= hs["lat"] <= lat_max and lon_min <= hs["lon"] <= lon_max:
                    hs_by_prov[pname] += 1
                    break

        return {
            "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_hotspots": len(hotspots),
            "sites_at_risk": len(site_fire),
            "critical_sites": sum(1 for s in site_fire if s["nearest_km"] < 1),
            "tickets_near_fire": len(near_tickets),
            "summary": {
                "total_hotspots": len(hotspots),
                "sites_at_risk": len(site_fire),
                "tickets_near_fire": len(near_tickets),
                "top_province": top_prov,
                "critical_sites": sum(1 for s in site_fire if s["nearest_km"] < 1),
                "high_risk_sites": sum(1 for s in site_fire if s["nearest_km"] < 3),
            },
            "sites": site_fire[:30],
            "hotspot_coords": [
                {"lat": hs["lat"], "lon": hs["lon"], "frp": hs["frp"],
                 "brightness": hs["brightness"], "confidence": hs["confidence"]}
                for hs in hotspots[:500]
            ],
            "by_province": [
                {"prov": k, "hotspots": v}
                for k, v in sorted(hs_by_prov.items(), key=lambda x: -x[1])
            ],
            "correlation": {
                "near_count": len(near_tickets),
                "far_count": len(far_tickets),
                "near_avg_aging": near_avg_aging,
                "far_avg_aging": far_avg_aging,
                "near_power_pct": near_power_pct,
                "far_power_pct": far_power_pct,
                "near_over_sla_pct": near_over_pct,
                "far_over_sla_pct": far_over_pct,
                "aging_ratio": round(near_avg_aging / far_avg_aging, 2) if far_avg_aging > 0 else 1,
                "power_ratio": round(near_power_pct / far_power_pct, 2) if far_power_pct > 0 else 1,
                "top_roots_near_fire": top_near_roots,
            },
            "prov_risk": [
                {"prov": k, **v}
                for k, v in sorted(prov_risk.items(), key=lambda x: -x[1]["hotspots_5km"])
            ],
        }

    def build_tab(rows):
        if not rows:
            return None

        total = len(rows)
        itime = max((r["itime"] for r in rows), default="")
        coords = [
            {
                "lat": r["lat"],
                "lon": r["lon"],
                "af": r["af"],
                "sev": r["sev"],
                "sowner": r["sowner"],
                "prov": r["prov"],
                "aging": r["aging"],
                "problem": r["problem"],
                "root": r["root"],
                "district": r["district"],
                "subimpact": r["subimpact"],
            }
            for r in rows
            if r["hc"]
        ]

        af_c = Counter(r["af"] for r in rows)
        sev_af = defaultdict(lambda: {k: 0 for k in AF_ORDER})
        sev_t = Counter()
        for r in rows:
            sev_af[r["sev"]][r["af"]] += 1
            sev_t[r["sev"]] += 1

        sev_af_l = [{"name": s, "total": sev_t[s], **sev_af[s]} for s in sorted(sev_t.keys(), key=lambda x: -sev_t[x])]

        sub_v = [r["subimpact"] for r in rows]
        sub_nz = [v for v in sub_v if v > 0]
        sub_sum = {"total": sum(sub_v), "count_nonzero": len(sub_nz)}

        dd = defaultdict(lambda: {"c": 0, "prov": "", "reg": "", "af": Counter(), "as": 0, "st": 0, "sc": 0, "ls": 0, "lo": 0, "cc": 0})
        for r in rows:
            dk = r["district"] or "ไม่ระบุ"
            d = dd[dk]
            d["c"] += 1
            d["prov"] = r["prov"] or d["prov"]
            d["reg"] = r["reg"]
            d["af"][r["af"]] += 1
            d["as"] += r["aging"]
            if r["subimpact"] > 0:
                d["st"] += r["subimpact"]
                d["sc"] += 1
            if r["hc"]:
                d["ls"] += r["lat"]
                d["lo"] += r["lon"]
                d["cc"] += 1

        districts = [
            {
                "name": dk,
                "count": d["c"],
                "prov": d["prov"],
                "reg": d["reg"],
                "avg": round(d["as"] / d["c"], 2),
                "sub_total": d["st"],
                "sub_count": d["sc"],
                "lat": round(d["ls"] / d["cc"], 4) if d["cc"] else 0,
                "lon": round(d["lo"] / d["cc"], 4) if d["cc"] else 0,
                **{k: d["af"].get(k, 0) for k in AF_ORDER},
            }
            for dk, d in sorted(dd.items(), key=lambda x: -x[1]["c"])
        ]

        od = defaultdict(lambda: {"c": 0, "reg": "", "prov": "", "af": Counter(), "as": 0, "sd": 0, "cd": 0, "pf": 0, "ip": 0, "st": 0, "sc": 0})
        for r in rows:
            d = od[r["sowner"]]
            d["c"] += 1
            d["reg"] = r["reg"]
            if not d["prov"] and r["prov"]:
                d["prov"] = r["prov"]
            d["af"][r["af"]] += 1
            d["as"] += r["aging"]
            if r["problem"] == "SITE DOWN":
                d["sd"] += 1
            elif r["problem"] == "CELL DOWN":
                d["cd"] += 1
            if r["root"] == "MAIN AC POWER FAIL":
                d["pf"] += 1
            if "IPRAN" in r["root"]:
                d["ip"] += 1
            if r["subimpact"] > 0:
                d["st"] += r["subimpact"]
                d["sc"] += 1

        owners = [
            {
                "name": o,
                "count": d["c"],
                "reg": d["reg"],
                "prov": d["prov"],
                "avg": round(d["as"] / d["c"], 2),
                "sd": d["sd"],
                "cd": d["cd"],
                "pf": d["pf"],
                "ip": d["ip"],
                "sub_total": d["st"],
                "sub_count": d["sc"],
                **{k: d["af"].get(k, 0) for k in AF_ORDER},
            }
            for o, d in sorted(od.items(), key=lambda x: -x[1]["c"])
        ]

        rs = {}
        for reg in ["NOR1", "NOR2"]:
            rr = [r for r in rows if r["reg"] == reg]
            if not rr:
                continue
            raf = Counter(r["af"] for r in rr)
            rs[reg] = {
                "total": len(rr),
                "avg_aging": round(sum(r["aging"] for r in rr) / len(rr), 2),
                **{k: raf.get(k, 0) for k in AF_ORDER},
            }

        rf = defaultdict(lambda: {k: 0 for k in AF_ORDER})
        rt = Counter()
        for r in rows:
            rf[r["root"]][r["af"]] += 1
            rt[r["root"]] += 1

        root_af = [{"name": rc, "total": rt[rc], **rf[rc]} for rc in sorted(rt.keys(), key=lambda x: -rt[x])]

        cat_c = Counter(r["cat"] for r in rows)

        need_team = []
        for d2 in districts:
            if d2["name"] == "ไม่ระบุ":
                continue
            need_team.append(
                {
                    "name": d2["name"],
                    "prov": d2["prov"],
                    "reg": d2["reg"],
                    "count": d2["count"],
                    "avg": d2["avg"],
                    "score": round(d2["count"] * d2["avg"], 1),
                    "over": d2["count"] - d2.get("6) Within SLA", 0),
                }
            )
        need_team.sort(key=lambda x: -x["score"])

        worst = max(owners, key=lambda x: x["avg"]) if owners else None
        best = min(owners, key=lambda x: x["avg"]) if owners else None
        over30 = af_c.get("1) OverSLA : > 30 days", 0)
        within = af_c.get("6) Within SLA", 0)
        n1 = rs.get("NOR1", {})
        n2 = rs.get("NOR2", {})
        td2 = next((d2 for d2 in districts if d2["name"] != "ไม่ระบุ"), None)
        weather = build_weather(rows)
        hotspots = fetch_firms_hotspots()
        fire = build_fire(rows, hotspots)
        tickets = [r["ticket"] for r in rows]
        geo_clusters = cluster_5km(rows)
        pri_c = Counter(r["pri"] for r in rows)

        return {
            "total": total,
            "itime": itime,
            "af": {k: af_c.get(k, 0) for k in AF_ORDER},
            "sev_af": sev_af_l,
            "reg_sum": rs,
            "owners": owners,
            "root_af": root_af,
            "districts": districts[:25],
            "subimpact": sub_sum,
            "sev": dict(Counter(r["sev"] for r in rows).most_common()),
            "cat": dict(cat_c.most_common(10)),
            "ven": dict(Counter(r["vendor"] for r in rows).most_common()),
            "prob": dict(Counter(r["problem"] for r in rows).most_common()),
            "aging_avg": round(sum(r["aging"] for r in rows) / total, 2),
            "aging_max": round(max(r["aging"] for r in rows), 2),
            "coords": coords[:500],
            "need_team": need_team[:5],
            "weather": weather,
            "fire": fire,
            "tickets": tickets,
            "clusters": geo_clusters,
            "priority": {"p0": pri_c.get(0, 0), "p1": pri_c.get(1, 0), "p2": pri_c.get(2, 0)},
            "insight": {
                "within_pct": round(within / total * 100, 1),
                "over30": over30,
                "over30_pct": round(over30 / total * 100, 1) if total else 0,
                "worst": worst["name"] if worst else "-",
                "worst_avg": worst["avg"] if worst else 0,
                "best": best["name"] if best else "-",
                "best_avg": best["avg"] if best else 0,
                "nor1_total": n1.get("total", 0),
                "nor1_over": n1.get("total", 0) - n1.get("6) Within SLA", 0),
                "nor2_total": n2.get("total", 0),
                "nor2_over": n2.get("total", 0) - n2.get("6) Within SLA", 0),
                "top_root": root_af[0]["name"] if root_af else "-",
                "top_root_count": root_af[0]["total"] if root_af else 0,
                "top_dist": td2["name"] if td2 else "-",
                "top_dist_count": td2["count"] if td2 else 0,
                "sub_total": sub_sum["total"],
                "sub_affected": sub_sum["count_nonzero"],
            },
        }

    tabs_def = [
        ("tab1", "📊 ภาพรวม Total", lambda r: True),
        ("tab2", "🔴 SA1-4", lambda r: r["sev"] in ("SA1", "SA2", "SA3", "SA4")),
        ("tab3", "📱 7.MB SA1-4", lambda r: r["bookmark"] == "7.MB with SA1-4"),
        ("tab4", "🌐 4.FBB SA1-4", lambda r: r["bookmark"] == "4.FBB with SA1-4"),
        ("tab5", "⚠️ NW Incident NSA1-2", lambda r: r["bookmark"] == "3. All NW Incident NSA1-2"),
        ("tab6", "📋 NSA3-4", lambda r: r["sev"] in ("NSA3", "NSA4")),
    ]

    result = {}
    for tid, name, filt in tabs_def:
        filtered = [r for r in all_rows if filt(r)]
        td = build_tab(filtered)
        if td:
            result[tid] = {"name": name, "data": td}
            print("  " + tid + ": " + str(td["total"]) + " tkts, " + str(len(td["clusters"])) + " clusters")

    return result


def gen_html(tabs_data, src):
    d_js = json.dumps(tabs_data, ensure_ascii=False)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not TPL_FILE.exists():
        raise FileNotFoundError(f"Template file not found: {TPL_FILE}")

    with open(TPL_FILE, "r", encoding="utf-8") as f:
        component = f.read()

    script_content = component.replace("__DATA_PLACEHOLDER__", d_js)

    html = '<!DOCTYPE html>\n<html lang="th"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">\n'
    html += '<title>BBTEC Intelligent Dashboard</title>\n'
    html += '<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Sarabun:wght@400;500;600;700&display=swap" rel="stylesheet">\n'
    html += '<script src="https://unpkg.com/react@18/umd/react.production.min.js"></script>\n'
    html += '<script src="https://unpkg.com/react-dom@18/umd/react-dom.production.min.js"></script>\n'
    html += '<script src="https://unpkg.com/@babel/standalone/babel.min.js"></script>\n'
    html += '<style>*{margin:0;padding:0;box-sizing:border-box}body{font-family:Inter,Sarabun,system-ui,sans-serif}button{font-family:inherit}::-webkit-scrollbar{width:6px}::-webkit-scrollbar-track{background:#1e2430}::-webkit-scrollbar-thumb{background:#3c4556;border-radius:3px}</style>\n'
    html += '</head><body><div id="root"></div>\n'
    html += '<div style="position:fixed;bottom:8px;left:8px;background:rgba(40,50,60,0.8);color:#8894a6;padding:4px 10px;border-radius:6px;font-size:9px;font-family:monospace;z-index:999">Updated: ' + ts + ' | ' + src + '</div>\n'
    html += '<script type="text/babel">\n'
    html += 'const{useState,useMemo,useEffect,useRef}=React;\n'
    html += script_content
    html += '\n</script></body></html>'
    return html


def get_last():
    if LAST_TRACKER.exists():
        return LAST_TRACKER.read_text(encoding="utf-8").strip()
    return ""


def set_last(s):
    LAST_TRACKER.write_text(s, encoding="utf-8")


def check():
    print("\n" + "=" * 50)
    print("Checking... (" + datetime.now().strftime("%H:%M:%S") + ")")
    try:
        fp, fn = find_best_file()
        if not fp:
            print("  No files found")
            return

        sig = fn + "_" + str(os.path.getmtime(fp))
        if sig == get_last():
            print("  Up to date: " + fn)
            return

        tabs_data = process_excel(fp)
        if not tabs_data:
            return

        OUTPUT_DIR.mkdir(exist_ok=True)
        html = gen_html(tabs_data, fn)
        out = OUTPUT_DIR / "dashboard.html"

        with open(out, "w", encoding="utf-8") as f:
            f.write(html)

        total = sum(t["data"]["total"] for t in tabs_data.values())

        summary = {
            "source_file": fn,
            "total_tickets": total,
            "tabs": len(tabs_data),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        summary_file = OUTPUT_DIR / "summary.json"
        with open(summary_file, "w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        set_last(sig)

        print("\n  Dashboard updated! -> " + str(out))
        print("  Summary updated! -> " + str(summary_file))
        print("  Total: " + str(total) + " tickets across " + str(len(tabs_data)) + " tabs")

    except Exception as e:
        print("  Error: " + str(e))
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    print("BBTEC Intelligent Dashboard — Auto-Updater")
    print("Output: " + str(OUTPUT_DIR))

    if not TPL_FILE.exists():
        print("ERROR: Template file not found: " + str(TPL_FILE))
        sys.exit(1)

    print("Template: " + str(TPL_FILE) + " (" + str(TPL_FILE.stat().st_size // 1024) + " KB)")
    check()
    print("\nFinished.\n")
